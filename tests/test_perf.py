"""
تست‌های آفلاینِ موتورِ متریکِ دقت و خروجیِ Excelِ خطاها (بدون API).

اجرا:  python -m pytest -q tests/test_perf.py
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from src.reporting import metrics as M  # noqa: E402
from src.reporting.errors_export import (  # noqa: E402
    jsonl_to_xlsx,
    read_errors_jsonl,
    write_errors_xlsx,
)

RES = {
    "n": 300, "errors": 0, "model": "deepseek-v4-pro", "latency_ms_avg": 742.0,
    "overall": {"accuracy": 0.893, "correct": 268, "total": 300},
    "confidence": {"confident_total": 250, "confident_correct": 240,
                   "flagged_total": 50, "flagged_correct": 28},
    "layers": [
        {"id": "layer1", "name": "Type", "accuracy": 0.91, "correct": 273, "total": 300,
         "label_ids": ["incident", "service_request"],
         "classes": [{"id": "incident", "name": "Incident", "recall": 0.94, "correct": 169, "total": 180},
                     {"id": "service_request", "name": "Service Request", "recall": 0.87, "correct": 104, "total": 120}],
         "confusion": {"incident": {"incident": 169, "service_request": 11},
                       "service_request": {"service_request": 104, "incident": 16}}},
        {"id": "layer2", "name": "Domain", "accuracy": 0.95, "correct": 285, "total": 300,
         "label_ids": ["erp", "staff"],
         "classes": [{"id": "erp", "name": "ERP", "recall": 0.96, "correct": 173, "total": 180},
                     {"id": "staff", "name": "Staff", "recall": 0.93, "correct": 112, "total": 120}],
         "confusion": {"erp": {"erp": 173, "staff": 7}, "staff": {"staff": 112, "erp": 8}}},
    ],
}


# ---------- متریک‌ها ----------
def test_precision_recall_f1_from_confusion():
    m = M.from_eval(RES, target=0.90)
    L = m.layers[0]
    inc = next(c for c in L.classes if c.id == "incident")
    # support = سطر (واقعی)، predicted = ستون (پیش‌بینی‌شده)
    assert inc.support == 180 and inc.tp == 169
    assert inc.predicted == 169 + 16  # ستونِ incident
    assert abs(inc.recall - 169 / 180) < 1e-12
    assert abs(inc.precision - 169 / 185) < 1e-12
    p, r = inc.precision, inc.recall
    assert abs(inc.f1 - 2 * p * r / (p + r)) < 1e-12


def test_macro_averages():
    m = M.from_eval(RES)
    L = m.layers[0]
    assert abs(L.macro_f1 - sum(c.f1 for c in L.classes) / 2) < 1e-12
    assert abs(L.macro_precision - sum(c.precision for c in L.classes) / 2) < 1e-12


def test_pass_fail_against_target():
    assert M.from_eval(RES, target=0.90).passed is False  # 0.893 < 0.90
    assert M.from_eval(RES, target=0.85).passed is True


def test_operational_readiness():
    r = M.from_eval(RES).readiness
    assert r.has_data
    assert abs(r.auto_coverage - 250 / 300) < 1e-12
    assert abs(r.auto_accuracy - 240 / 250) < 1e-12
    assert abs(r.review_share - 50 / 300) < 1e-12
    assert abs(r.review_accuracy - 28 / 50) < 1e-12


def test_readiness_absent_when_no_confidence():
    res = {**RES}
    res.pop("confidence")
    assert M.from_eval(res).readiness.has_data is False


# ---------- خروجیِ Excel ----------
_ERR = [
    {"Key": "INC-1", "wrong": ["layer1"], "Summary": "خطا در ثبت پانچ", "Description": "ثبت نشد",
     "layer1": {"true": "incident", "pred": "service_request"}, "layer2": {"true": "erp", "pred": "erp"}},
    {"Key": "SR-2", "wrong": ["layer2"], "Summary": "وام", "Description": "فعال نیست",
     "layer1": {"true": "service_request", "pred": "service_request"}, "layer2": {"true": "staff", "pred": "erp"}},
]


def test_write_errors_xlsx_headers_and_rows(tmp_path):
    import openpyxl

    path = write_errors_xlsx(_ERR, tmp_path / "errors.xlsx")
    assert path.exists()
    ws = openpyxl.load_workbook(path).active
    headers = [c.value for c in ws[1]]
    assert headers == ["Key", "wrong", "Summary", "Description",
                       "layer1_true", "layer1_pred", "layer2_true", "layer2_pred"]
    assert ws.max_row == 3                 # هدر + ۲ ردیف
    assert ws["A2"].value == "INC-1"
    assert ws["F2"].value == "service_request"  # layer1_pred
    assert ws.freeze_panes == "A2"


def test_jsonl_to_xlsx_roundtrip(tmp_path):
    import openpyxl

    jl = tmp_path / "errs.jsonl"
    jl.write_text("\n".join(json.dumps(r, ensure_ascii=False) for r in _ERR), encoding="utf-8")
    assert read_errors_jsonl(jl) == _ERR
    xp = jsonl_to_xlsx(jl, tmp_path / "errs.xlsx")
    assert openpyxl.load_workbook(xp).active.max_row == 3


def test_write_errors_xlsx_empty(tmp_path):
    import openpyxl

    path = write_errors_xlsx([], tmp_path / "empty.xlsx")  # نباید crash کند
    ws = openpyxl.load_workbook(path).active
    assert [c.value for c in ws[1]] == ["Key", "wrong", "Summary", "Description"]
    assert ws.max_row == 1
