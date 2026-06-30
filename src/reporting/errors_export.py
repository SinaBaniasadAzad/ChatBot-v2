"""
خروجیِ Excel از تیکت‌های اشتباه (مکملِ همان JSONL که eval می‌سازد).

رکوردهای خطا (از `scripts.eval_incdb`) به این شکل‌اند:
    {"Key": "...", "wrong": ["layer1"], "Summary": "...", "Description": "...",
     "layer1": {"true": "...", "pred": "..."}, "layer2": {...}}

این ماژول آن‌ها را به یک جدولِ تخت و خوانا در .xlsx تبدیل می‌کند: هدرِ پررنگ و
ثابت، AutoFilter، پهنای ستونِ مناسب، و پیچشِ متن برای ستون‌های بلند.
نکتهٔ PII: این فایل متنِ خامِ تیکت دارد؛ مثل JSONL باید محرمانه نگه داشته شود.
"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_WRAP = Alignment(wrap_text=True, vertical="top")
_TOP = Alignment(vertical="top")

# پهنای ستون‌ها (نام ستون -> عرض). پیش‌فرض برای ستون‌های ناشناخته.
_WIDTHS = {"Key": 14, "wrong": 14, "Summary": 46, "Description": 64}
_DEFAULT_WIDTH = 18


def _layer_keys(records: list[dict]) -> list[str]:
    """کلیدهای لایه را (به ترتیبِ ظهور) پیدا می‌کند: هر کلیدی که مقدارش {true,pred} است."""
    seen: list[str] = []
    for r in records:
        for k, v in r.items():
            if isinstance(v, dict) and "true" in v and "pred" in v and k not in seen:
                seen.append(k)
    return seen


def write_errors_xlsx(records: list[dict], path: str | Path) -> Path:
    """جدولِ تیکت‌های اشتباه را در یک فایلِ .xlsx می‌نویسد و مسیر را برمی‌گرداند."""
    layers = _layer_keys(records)
    headers = ["Key", "wrong", "Summary", "Description"]
    for lk in layers:
        headers += [f"{lk}_true", f"{lk}_pred"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Misclassified"

    ws.append(headers)
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _TOP

    for r in records:
        row = [
            r.get("Key", ""),
            ", ".join(r.get("wrong", []) or []),
            r.get("Summary", ""),
            r.get("Description", ""),
        ]
        for lk in layers:
            cell = r.get(lk) or {}
            row += [cell.get("true", ""), cell.get("pred", "")]
        ws.append(row)

    # استایل: پیچشِ متن برای Summary/Description، بالاترازِ بقیه
    for col_idx, name in enumerate(headers, 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = _WIDTHS.get(name, _DEFAULT_WIDTH)
        wrap = name in ("Summary", "Description")
        for cell in ws[letter][1:]:  # به‌جز هدر
            cell.alignment = _WRAP if wrap else _TOP

    ws.freeze_panes = "A2"
    if ws.max_row >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def read_errors_jsonl(path: str | Path) -> list[dict]:
    """رکوردهای خطا را از فایلِ JSONL می‌خواند (خطوطِ خراب نادیده گرفته می‌شوند)."""
    records: list[dict] = []
    for line in Path(path).read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            records.append(json.loads(line))
        except json.JSONDecodeError:
            continue
    return records


def jsonl_to_xlsx(jsonl_path: str | Path, xlsx_path: str | Path) -> Path:
    """تبدیلِ فایلِ JSONLِ خطا به .xlsx."""
    return write_errors_xlsx(read_errors_jsonl(jsonl_path), xlsx_path)
